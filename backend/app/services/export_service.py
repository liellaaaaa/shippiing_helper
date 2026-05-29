"""Excel 导出服务 — FR-5.x 数据看板模块"""

import io
from datetime import datetime
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Border, Side, Alignment


class ExportService:
    """Excel 导出服务"""

    HEADER_FILL = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
    HEADER_FONT = Font(bold=True, color="FFFFFF")
    BORDER_SIDE = Side(style="thin", color="000000")
    BORDER = Border(
        left=BORDER_SIDE,
        right=BORDER_SIDE,
        top=BORDER_SIDE,
        bottom=BORDER_SIDE,
    )

    # 列定义
    COLUMNS = [
        ("订单号", "order_no"),
        ("客户编码", "customer_code"),
        ("业务员", "salesperson"),
        ("内部编码", "internal_code"),
        ("产品名称", "product_cn"),
        ("订单数量", "order_quantity"),
        ("PI 数量", "pi_quantity"),
        ("差异状态", "diff_status"),
        ("关联状态", "association_status"),
    ]

    def generate_excel_bytes(self, data: list[dict]) -> bytes:
        """
        将合并数据生成为 Excel 文件流。
        返回 bytes，可在 FastAPI 中直接作为 Response。
        """
        wb = Workbook()
        ws = wb.active
        ws.title = "数据看板"

        # 写入表头
        headers = [col[0] for col in self.COLUMNS]
        for col_idx, header in enumerate(headers, start=1):
            cell = ws.cell(row=1, column=col_idx, value=header)
            cell.font = self.HEADER_FONT
            cell.fill = self.HEADER_FILL
            cell.border = self.BORDER
            cell.alignment = Alignment(horizontal="center", vertical="center")

        # 写入数据
        for row_idx, row_data in enumerate(data, start=2):
            for col_idx, (_, field_key) in enumerate(self.COLUMNS, start=1):
                value = row_data.get(field_key, "")
                cell = ws.cell(row=row_idx, column=col_idx, value=value)
                cell.border = self.BORDER

        # 冻结第一行
        ws.freeze_panes = "A2"

        # 列宽自适应
        for col_idx, (_, _) in enumerate(self.COLUMNS, start=1):
            max_length = len(headers[col_idx - 1])
            for row_idx in range(2, len(data) + 2):
                cell_value = ws.cell(row=row_idx, column=col_idx).value
                if cell_value:
                    max_length = max(max_length, len(str(cell_value)))
            ws.column_dimensions[chr(64 + col_idx)].width = min(max_length + 2, 30)

        # 转换为 bytes
        buffer = io.BytesIO()
        wb.save(buffer)
        buffer.seek(0)
        return buffer.getvalue()

    def generate_filename(self) -> str:
        """生成下载文件名"""
        today = datetime.now().strftime("%Y%m%d")
        return f"数据看板_{today}.xlsx"
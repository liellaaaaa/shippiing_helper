"""
PI文件解析器 - 从Excel文件中提取PI合同数据

支持两种格式：
1. 表格格式（外贸订单PI合同.xlsx）：标准行列格式，有表头
2. Proforma Invoice 格式（.xls/.xlsx）：自由文本型发票，含产品表格和银行信息
"""

import re
import io
import os
import json
import math
import pymupdf
import pytesseract
from PIL import Image
from openpyxl import load_workbook
import xlrd
from typing import Optional

from app.schemas.pi_contract import (
    PiContractUploadResponse,
    PiContractItemRow,
    ConfidenceInfo,
)


# 标准字段映射：字段名 -> 可能的中英文表头别名
COLUMN_MAPPING: dict[str, list[str]] = {
    "customer_code": ["客户编码", "客户编号", "客户代码", "Customer Code", "customer"],
    "pi_no": ["PI号", "PI NO.", "Proforma Invoice No.", "PI号码", "PI_NO", "Proforma Invoice Number", "发票号"],
    "sales_person": ["业务员", "Salesperson", "销售员"],
    "pi_date": ["日期", "Date", "PI Date", "开票日期"],
    "order_id": ["销售订单号", "Sales Order No.", "SO No.", "订单号"],
    "internal_code": ["内部编码", "Item Code", "SKU", "产品代码", "编码"],
    "quantity": ["数量", "QTY", "Quantity", "订单数量"],
    "unit_price": ["单价", "Unit Price", "Price", "单价(USD)"],
    "total_amount": ["金额", "Amount", "Total", "总额"],
    "product_color": ["产品颜色", "Color", "颜色"],
    "hs_code": ["海关编码", "H.S. Code", "HS Code", "编码"],
    "customs_name": ["报关品名", "Customs Name", "品名"],
    "customs_composition": ["报关成分", "Ingredients", "成分"],
    "order_customs_name": ["填写订单报关品名", "Order Customs Name"],
    "is_ordered": ["是否下单", "Is Ordered", "已下单"],
    "notes": ["备注", "Notes", "Remark", "备注信息"],
}


def _normalize_column_name(col_name: str) -> str | None:
    """
    将列名标准化为目标字段名

    - 去除空格
    - 匹配 COLUMN_MAPPING 中的别名
    - 返回字段名或 None（无法识别）
    """
    col_name = col_name.strip()
    for field_name, aliases in COLUMN_MAPPING.items():
        if col_name in aliases:
            return field_name
    return None


def _parse_float(value: str) -> float | None:
    """
    安全地解析浮点数

    - 空字符串返回 None
    - 解析失败返回 None
    """
    if not value or not value.strip():
        return None
    try:
        return float(value.strip())
    except (ValueError, TypeError):
        return None


def _build_header_map(header_row: list[str]) -> dict[int, str]:
    """
    构建列索引到目标字段名的映射

    根据表头行，返回 {列索引: 字段名} 的字典
    """
    col_map: dict[int, str] = {}
    for idx, col_name in enumerate(header_row):
        field_name = _normalize_column_name(col_name)
        if field_name:
            col_map[idx] = field_name
    return col_map


def parse_pi_file(rows: list[list[str]]) -> PiContractUploadResponse:
    """
    解析表格格式PI文件数据（标准行列格式）

    参数:
        rows: Excel行列表，第一行是表头

    返回:
        PiContractUploadResponse

    异常:
        ValueError: 表头为空或缺少关键字段（pi_no, customer_code）
    """
    if not rows or len(rows) < 1:
        raise ValueError("文件为空或无有效数据")

    header_row = rows[0]
    if not header_row or all(not c.strip() for c in header_row):
        raise ValueError("表头行无效")

    col_map = _build_header_map(header_row)

    # 关键字段检查：pi_no 和 customer_code
    if "pi_no" not in col_map.values() or "customer_code" not in col_map.values():
        raise ValueError("缺少关键字段：pi_no 或 customer_code")

    data_rows = rows[1:]

    # 构建反向映射：字段名 -> 列索引
    field_to_col: dict[str, int] = {}
    for col_idx, field_name in col_map.items():
        field_to_col[field_name] = col_idx

    # 公共字段（取第一行的值）
    pi_no = None
    customer_code = None
    sales_person = None
    pi_date = None
    is_ordered = "0"

    if data_rows:
        first_row = data_rows[0]
        for field_name, col_idx in field_to_col.items():
            if col_idx < len(first_row):
                val = first_row[col_idx].strip() if first_row[col_idx] else None
                if field_name == "pi_no":
                    pi_no = val
                elif field_name == "customer_code":
                    customer_code = val
                elif field_name == "sales_person":
                    sales_person = val
                elif field_name == "pi_date":
                    pi_date = val
                elif field_name == "is_ordered":
                    is_ordered = val if val else "0"

    # 解析明细行
    items: list[PiContractItemRow] = []
    row_errors: list[str] = []
    failed_rows = 0

    for row_idx, row in enumerate(data_rows):
        item_row_errors: list[str] = []
        missing_fields: list[str] = []

        internal_code = None
        quantity = None
        unit_price = None
        total_amount = None
        product_color = None
        hs_code = None
        customs_name = None
        customs_composition = None
        order_customs_name = None
        notes = None

        for field_name, col_idx in field_to_col.items():
            if col_idx < len(row):
                raw_val = row[col_idx]
                if field_name in ("quantity", "unit_price", "total_amount"):
                    parsed = _parse_float(raw_val)
                    if raw_val and raw_val.strip() and parsed is None:
                        item_row_errors.append(f"{field_name}解析失败: {raw_val}")
                    if field_name == "quantity":
                        quantity = parsed
                    elif field_name == "unit_price":
                        unit_price = parsed
                    elif field_name == "total_amount":
                        total_amount = parsed
                else:
                    val = raw_val.strip() if raw_val else None
                    if field_name == "internal_code":
                        internal_code = val
                        if not val:
                            missing_fields.append("internal_code")
                    elif field_name == "product_color":
                        product_color = val
                    elif field_name == "hs_code":
                        hs_code = val
                    elif field_name == "customs_name":
                        customs_name = val
                    elif field_name == "customs_composition":
                        customs_composition = val
                    elif field_name == "order_customs_name":
                        order_customs_name = val
                    elif field_name == "notes":
                        notes = val

        row_status = "success"
        row_error_msg = None
        if item_row_errors or missing_fields:
            row_status = "error"
            failed_rows += 1
            error_parts = item_row_errors.copy()
            if missing_fields:
                error_parts.append(f"缺少字段: {', '.join(missing_fields)}")
            row_error_msg = "; ".join(error_parts)
            row_errors.append(row_error_msg)

        items.append(PiContractItemRow(
            row_index=row_idx + 1,
            status=row_status,
            error_msg=row_error_msg,
            internal_code=internal_code,
            quantity=quantity,
            unit_price=unit_price,
            total_amount=total_amount,
            product_color=product_color,
            hs_code=hs_code,
            customs_name=customs_name,
            customs_composition=customs_composition,
            order_customs_name=order_customs_name,
            notes=notes,
        ))

    recognized = len(col_map)
    total_cols = len(header_row)
    confidence = ConfidenceInfo(
        recognized=recognized,
        total=total_cols,
        percentage=f"{recognized / total_cols * 100:.0f}%" if total_cols > 0 else "0%",
        failed_rows=failed_rows,
    )

    if not pi_no:
        raise ValueError("无法从数据中提取 PI号")

    return PiContractUploadResponse(
        pi_no=pi_no or "",
        customer_code=customer_code,
        sales_person=sales_person,
        pi_date=pi_date,
        is_ordered=is_ordered or "0",
        items=items,
        confidence=confidence,
    )


# ──────────────────────────────────────────────────────────────────────
# Proforma Invoice 解析辅助函数
# ──────────────────────────────────────────────────────────────────────

def _clean_number(value: str) -> float | None:
    """清洗并解析数值字段，移除 ¥ , 等符号"""
    if not value:
        return None
    value = value.replace("¥", "").replace(",", "").replace(" ", "").strip()
    try:
        return float(value)
    except (ValueError, TypeError):
        return None


def _excel_serial_to_date(serial: float) -> str | None:
    """将 Excel 日期序列号转换为 YYYY-MM-DD 格式"""
    try:
        from datetime import datetime, timedelta
        base = datetime(1899, 12, 30)
        return (base + timedelta(days=int(serial))).strftime("%Y-%m-%d")
    except (ValueError, TypeError):
        return None


def _normalize_hs_code(raw: str | None) -> str | None:
    """
    标准化 H.S. Code：
    - 去除所有空格和点号
    - 重新格式化为 XXXX.XX.XX.XX 形式（如果长度合适）
    """
    if not raw:
        return None
    # 去除空格、点号、制表符
    cleaned = re.sub(r'[\s.\t]', '', raw.strip())
    if not cleaned or not cleaned[0].isdigit():
        return None
    # 尝试重新格式化：如果纯数字，按标准格式加点
    if cleaned.isdigit() and len(cleaned) >= 6:
        # H.S. code 格式：前4位 + 后续2位一组
        parts = []
        for i in range(0, len(cleaned), 2):
            parts.append(cleaned[i:i+2])
        return ".".join(parts)
    return cleaned


def _extract_field_from_rows(rows: list[list[str]], patterns: list[str]) -> str | None:
    """
    从行列结构中直接查找 Label→Value 模式（跨行、同一列）。
    """
    for row_idx, row in enumerate(rows):
        for col_idx, cell in enumerate(row):
            cell_str = str(cell).strip()
            for pattern in patterns:
                if re.search(rf'(?i){pattern}', cell_str):
                    # 优先返回当前格 after-colon 的值
                    current_val = cell_str.split(":")[-1].strip()
                    if current_val:
                        return current_val
                    # 当前格 after-colon 为空，检查下一行同列
                    if row_idx + 1 < len(rows):
                        next_row = rows[row_idx + 1]
                        if col_idx < len(next_row):
                            next_val = str(next_row[col_idx]).strip()
                            if next_val and not re.match(r'^[A-Za-z\s]+:\s*$', next_val):
                                return next_val
    return None


def _scan_all_cells_for_ht(rows: list[list[str]]) -> str | None:
    """
    遍历所有单元格，找到第一个匹配 HT 格式的订单号。
    支持 HT + 8位或10位日期 + 后缀。
    """
    for row in rows:
        for cell in row:
            cell_str = str(cell).strip()
            m = re.search(r'HT\d{6,10}[A-Z0-9]*', cell_str)
            if m:
                return m.group(0)
    return None


def _scan_pi_number(rows: list[list[str]], filename: str = "") -> str | None:
    """
    多策略扫描 PI 编号：
    1. 搜索 "PI No" / "Order No" / "PO No" 标签，取同行或下行的 HT 号
    2. 遍历所有单元格找 HT 格式的订单号
    3. 从文件名提取 HT 号
    """
    # 策略1: 找标签 "PI No" / "Order No" / "PO No"
    label_patterns = [
        r"PI\s*No", r"PI\s*Number", r"PROFORMA\s*INVOICE\s*NO",
        r"Order\s*No", r"ORDER\s*NO", r"ORDER\s*NUMBER",
        r"PO\s*No", r"PO\s*NUMBER",
    ]
    for ri, row in enumerate(rows):
        for ci, cell in enumerate(row):
            cell_str = str(cell).strip()
            for pat in label_patterns:
                if re.search(rf'(?i){pat}', cell_str):
                    # 从当前单元格中提取 HT 号
                    ht_m = re.search(r'HT\d{6,10}[A-Z0-9]*', cell_str)
                    if ht_m:
                        return ht_m.group(0)
                    # 检查同行其他单元格
                    for other_cell in row:
                        other_str = str(other_cell).strip()
                        ht_m = re.search(r'HT\d{6,10}[A-Z0-9]*', other_str)
                        if ht_m:
                            return ht_m.group(0)
                    # 检查下一行
                    if ri + 1 < len(rows):
                        for next_cell in rows[ri + 1]:
                            next_str = str(next_cell).strip()
                            ht_m = re.search(r'HT\d{6,10}[A-Z0-9]*', next_str)
                            if ht_m:
                                return ht_m.group(0)

    # 策略2: 扫描所有单元格找 HT 格式号
    ht_from_cells = _scan_all_cells_for_ht(rows)
    if ht_from_cells:
        return ht_from_cells

    # 策略3: 从文件名提取
    if filename:
        fn_pi = re.search(r'HT\d{6,10}[A-Z0-9]*', os.path.basename(filename))
        if fn_pi:
            return fn_pi.group(0)

    return None


def _scan_date(rows: list[list[str]]) -> str | None:
    """
    扫描日期字段：
    1. 找 "Date" 标签，取同行或下行的值
    2. 支持 Excel 序列号和日期字符串
    """
    for ri, row in enumerate(rows):
        for ci, cell in enumerate(row):
            cell_str = str(cell).strip()
            if re.search(r'(?i)^Date\s*[:.]?\s*$', cell_str) or re.search(r'(?i)^Date$', cell_str):
                # 同行其他列
                for other in row:
                    other_str = str(other).strip()
                    if other_str and other_str.lower() != 'date':
                        # 可能是 Excel 序列号
                        try:
                            serial = float(other_str)
                            if serial > 30000:
                                return _excel_serial_to_date(serial)
                        except (ValueError, TypeError):
                            pass
                        if re.match(r'\d{4}[-/.]\d{1,2}[-/.]\d{1,2}', other_str):
                            return other_str
                # 下一行
                if ri + 1 < len(rows) and ci < len(rows[ri + 1]):
                    next_val = str(rows[ri + 1][ci]).strip()
                    try:
                        serial = float(next_val)
                        if serial > 30000:
                            return _excel_serial_to_date(serial)
                    except (ValueError, TypeError):
                        pass
                    if re.match(r'\d{4}[-/.]\d{1,2}[-/.]\d{1,2}', next_val):
                        return next_val

    # 兜底：扫描所有单元格找日期格式或 Excel 序列号（>30000）
    for row in rows:
        for cell in row:
            cell_str = str(cell).strip()
            # 日期字符串
            m = re.match(r'^(\d{4}[-/.]\d{1,2}[-/.]\d{1,2})', cell_str)
            if m:
                return m.group(1)
            # Excel 序列号（在 Date 标签附近出现的数字）
            try:
                serial = float(cell_str)
                if 30000 < serial < 50000:
                    return _excel_serial_to_date(serial)
            except (ValueError, TypeError):
                pass

    return None


def _scan_consignee(rows: list[list[str]]) -> tuple[str | None, str | None, str | None]:
    """
    扫描 Consignee 信息：Name, Address, Tel
    返回 (name, address, tel)
    """
    consignee_start = -1
    for ri, row in enumerate(rows):
        for cell in row:
            if re.search(r'(?i)Consignee\s*:', str(cell).strip()):
                consignee_start = ri
                break
        if consignee_start >= 0:
            break

    name = None
    address = None
    tel = None

    if consignee_start < 0:
        return name, address, tel

    # 在 Consignee 区域内搜索 Name 和 Address
    for ri in range(consignee_start, min(consignee_start + 6, len(rows))):
        row = rows[ri]
        for ci, cell in enumerate(row):
            cell_str = str(cell).strip()
            # Name
            m = re.match(r'(?i)Name\s*:\s*(.+)', cell_str)
            if m:
                name = m.group(1).strip()
            # Address
            m = re.match(r'(?i)(?:Add\.?|Address)\s*:\s*(.+)', cell_str, re.DOTALL)
            if m:
                address = m.group(1).strip()
            # Tel
            m = re.match(r'(?i)Tel\s*:\s*(.+)', cell_str)
            if m:
                tel = m.group(1).strip()

    # 如果 Name 在单元格里但 Address 在下一行（跨行布局）
    if name and not address:
        for ri in range(consignee_start, min(consignee_start + 6, len(rows))):
            row = rows[ri]
            for cell in row:
                cell_str = str(cell).strip()
                if re.match(r'(?i)(?:Add\.?|Address)\s*:', cell_str):
                    m = re.match(r'(?i)(?:Add\.?|Address)\s*:\s*(.+)', cell_str, re.DOTALL)
                    if m:
                        address = m.group(1).strip()
                    break
            if address:
                break

    # 如果 Name 和 Address 在同一单元格的 Name: 后面（如 "Name: XXX\nAddress: YYY"）
    if name and not address:
        for ri in range(consignee_start, min(consignee_start + 6, len(rows))):
            for cell in rows[ri]:
                cell_str = str(cell).strip()
                if 'Name' in cell_str and 'Address' in cell_str:
                    # 提取 Address 部分
                    addr_m = re.search(r'(?i)Address\s*:\s*(.+)', cell_str, re.DOTALL)
                    if addr_m:
                        address = addr_m.group(1).strip()
                    break
            if address:
                break

    # 如果 Name 在单元格中但后面没有 Address，检查 Name 所在单元格的完整内容
    # 有些文件 Name 和 Address 在同一个单元格的不同行
    if name and not address:
        for ri in range(consignee_start, min(consignee_start + 6, len(rows))):
            for cell in rows[ri]:
                cell_str = str(cell).strip()
                if name in cell_str and '\n' in cell_str:
                    lines = cell_str.split('\n')
                    for line in lines[1:]:
                        line = line.strip()
                        if line and not re.match(r'(?i)^(Name|Consignee|Date|PI)', line):
                            address = line
                            break
                if address:
                    break
            if address:
                break

    return name, address, tel


def _is_packing_row(row_text: str) -> bool:
    """
    检测是否是 PACKING 说明行（非产品行）。
    例如: "PACKING : 7X 1000KG" 或 "1. Packing 25kg/bag export standard packing"
    """
    text_upper = row_text.upper().strip()
    # "PACKING :" 开头
    if re.match(r'PACKING\s*[:：]', text_upper):
        return True
    # "X KG" / "X DRUM" 等包装规格
    if re.match(r'\d+\s*[Xx]\s*\d+', text_upper):
        return True
    # 数字 + KG/DRUM/BAG 等包装单位
    if re.match(r'[\d,.]+\s*(?:KG|DRUM|BAG|IBC|PCS|SET)', text_upper):
        # 但如果同时有单价和金额，可能是产品行
        numbers = re.findall(r'[\d,]+\.?\d*', row_text)
        if len(numbers) >= 3:
            return False  # 有3个以上数字，可能是产品行
        return True
    return False


def _is_remark_or_terms_row(row_text: str) -> bool:
    """检测是否是备注或条款行"""
    text_upper = row_text.upper().strip()
    if 'REMARKS' in text_upper:
        return True
    if re.match(r'\d+\.\s+(?:All|This|The|Quantity|Deliver|Packing\s+\d)', text_upper):
        return True
    return False


def _is_product_description_row(row_text: str) -> bool:
    """
    检测是否可能是产品描述行（文字为主，含产品名）。
    用于识别跨行产品描述（上一行品名，下一行数字）。
    """
    text = row_text.strip()
    if not text:
        return False
    # 纯数字行不是产品描述
    if re.match(r'^[\d,.\s]+$', text):
        return False
    # TOTAL 行不是
    if re.match(r'(?i)^TOTAL', text):
        return False
    # 包含英文字母（产品名通常是英文）
    if re.search(r'[A-Za-z]{2,}', text):
        return True
    # 包含中文字符（产品名可能是中文）
    if re.search(r'[\u4e00-\u9fff]', text):
        return True
    return False


def _extract_price_term_from_rows(rows: list[list[str]]) -> str | None:
    """
    从行结构中提取完整的价格条款。
    优先从 "Price Term:" 标签提取，其次从表头中提取（如 "CIF Jakarta(USD/Kg)"）。
    """
    # 策略1: 找 "Price Term:" 标签
    for ri, row in enumerate(rows):
        for cell in row:
            cell_str = str(cell).strip()
            m = re.search(r'(?i)Price\s+Term\s*[:：]\s*(.+)', cell_str)
            if m:
                term = m.group(1).strip()
                if term:
                    return term

    # 策略2: 从表头中提取价格条款（如 "CIF Jakarta(USD/Kg)" 或 "CIF KLANG( USD/kg)"）
    for row in rows:
        for cell in row:
            cell_str = str(cell).strip()
            m = re.search(
                r'\b(CIF|FOB|CFR|C\s*&\s*F|EXW|FCA|CPT|CIP|DAP|DPU)\b'
                r'\s*(.+?)(?:\s*[\(（]|$)',
                cell_str, re.IGNORECASE
            )
            if m:
                term_type = m.group(1).upper()
                location = m.group(2).strip()
                if location:
                    return f"{term_type} {location}"
                return term_type

    # 策略3: 兜底，只找 CIF/FOB 等关键字
    for row in rows:
        for cell in row:
            cell_str = str(cell).strip()
            m = re.search(r'\b(CIF|FOB|CFR|C\s*&\s*F|EXW|FCA|CPT|CIP|DAP|DPU)\b', cell_str, re.IGNORECASE)
            if m:
                return m.group(1).upper()

    return None


def _extract_currency_from_rows(rows: list[list[str]]) -> str | None:
    """
    从 PI 文件行中提取币制（USD / CNY / RMB）。
    策略：
    1. 从 Total 行提取（如 "TOTAL : 18000.00 US$49,720.00" 或 "Total Amount US$49,720.00"）
    2. 从表头提取（如 "单价(USD)"、"Amount (USD)"）
    3. 从价格条款提取（如 "CIF Jakarta(USD/Kg)"）
    """
    # 策略1: 从 Total/Amount 行提取
    for row in rows:
        for cell in row:
            cell_str = str(cell).strip()
            # US$ / USD / $
            m = re.search(r'(?i)(?:US\s*\$|USD|CNY|RMB|人民币|美元)', cell_str)
            if m:
                token = m.group(0).upper().replace("US$", "USD").replace("US $", "USD").replace("人民币", "CNY").replace("美元", "USD")
                if token in ("USD", "CNY", "RMB"):
                    return "USD" if token == "RMB" else token

    # 策略2: 从表头提取
    for row in rows[:5]:
        for cell in row:
            cell_str = str(cell).strip()
            m = re.search(r'(?i)\(?\s*(USD|CNY|RMB)\s*\)?', cell_str)
            if m:
                token = m.group(1).upper()
                return "USD" if token == "RMB" else token

    # 策略3: 从价格条款提取（如 "CIF Jakarta(USD/Kg)"）
    for row in rows[:5]:
        for cell in row:
            cell_str = str(cell).strip()
            m = re.search(r'(?i)(?:CIF|FOB|CFR|C\s*&\s*F)\s*.+?\(\s*(USD|CNY|RMB)', cell_str)
            if m:
                token = m.group(1).upper()
                return "USD" if token == "RMB" else token

    return None


def _find_hs_code_column(rows: list[list[str]]) -> int | None:
    """
    查找包含 HS CODE 的表头列索引。
    返回列索引或 None。
    """
    for row in rows[:5]:  # 只检查前5行
        for ci, cell in enumerate(row):
            cell_str = str(cell).strip()
            if re.match(r'(?i)^H\.?S\.?\s*(?:CODE|编码)?$', cell_str):
                return ci
    return None


def _extract_hs_code_from_rows(rows: list[list[str]]) -> str | None:
    """
    从行结构中提取 H.S. Code。
    支持 "H.S. Code:" 标签和独立列中的 HS code。
    """
    for ri, row in enumerate(rows):
        for cell in row:
            cell_str = str(cell).strip()
            m = re.search(r'(?i)H\.?S\.?\s*(?:Code|编码)\s*[:：]\s*(.+)', cell_str)
            if m:
                raw = m.group(1).strip()
                # HS code 可能紧跟其他文字（如 "3402.90.99" 或 "3402.90.99  125kg/drum"）
                hs_m = re.match(r'([\d\s.]+)', raw)
                if hs_m:
                    return _normalize_hs_code(hs_m.group(1))
                return _normalize_hs_code(raw)
    return None


def _extract_hs_from_description(desc: str) -> tuple[str | None, str]:
    """
    从产品描述中提取嵌入的 HS code。
    例如: "FIXING HT-7212\nHS code： 3402.90.99" → ("3402.90.99", "FIXING HT-7212")
    """
    m = re.search(r'(?i)HS\s*(?:code|编码)\s*[:：]\s*([\d\s.]+)', desc)
    if m:
        hs_code = m.group(1).strip()
        # 从描述中移除 HS code 部分
        clean_desc = re.sub(r'(?i)\n?\s*HS\s*(?:code|编码)\s*[:：]\s*[\d\s.]+', '', desc).strip()
        return hs_code, clean_desc
    return None, desc


def parse_proforma_invoice(rows: list[list[str]]) -> PiContractUploadResponse:
    """
    解析 Proforma Invoice 格式文档（非表格型发票）。

    支持多种变体：
    - PI No / Order No / PO No 标签
    - Consignee 区域（Name / Address / Tel）
    - 产品表格（DESCRIPTION OF GOODS）含 PACKING 列
    - 付款条款和银行信息
    - 价格条款从表头或标签提取

    适用文件：所有 Proforma Invoice 格式的 .xls/.xlsx 文件
    """

    # ── 1. PI 编号 ──────────────────────────────────────────────────
    pi_no = _scan_pi_number(rows)

    # ── 2. 日期 ──────────────────────────────────────────────────────
    pi_date = _scan_date(rows)

    # ── 3. Consignee 信息 ──────────────────────────────────────────
    consignee_name, consignee_address, consignee_tel = _scan_consignee(rows)

    # ── 4. 价格条款 ──────────────────────────────────────────────────
    price_term = _extract_price_term_from_rows(rows)

    # ── 5. H.S. Code（从标签提取） ──────────────────────────────────
    hs_code_from_label = _extract_hs_code_from_rows(rows)

    # ── 5b. 币制提取 ──────────────────────────────────────────────────
    currency = _extract_currency_from_rows(rows)

    # ── 6. 从 Payment terms and conditions 区域提取字段 ──────────
    payment_terms = None
    destination = None
    beneficiary_bank = None
    loading_port = None
    invoice_to = None
    in_payment_section = False

    for ri, row in enumerate(rows):
        for cell in row:
            cell_str = str(cell).strip()
            # 检测 payment section 开始：
            # 1) 显式 "Payment terms and conditions" 标题
            # 2) 编号列表中的 H.S. Code / Payment Terms / Destination（无标题的变体）
            if re.search(r'(?i)Payment\s+terms\s+and\s+conditions', cell_str):
                in_payment_section = True
            elif re.search(r'(?i)^\d+\.\s*H\.?S\.?\s*(?:Code|编码)\s*[:：]', cell_str):
                in_payment_section = True
            elif re.search(r'(?i)^\d+\.\s*Payment\s+Terms?\s*[:：]', cell_str):
                in_payment_section = True
            elif re.search(r'(?i)^\d+\.\s*Destination\s*[:：]', cell_str):
                in_payment_section = True

            if in_payment_section:
                m = re.search(r'(?i)Payment\s+Terms?\s*[:：]\s*(.+)', cell_str)
                if m:
                    payment_terms = m.group(1).strip()
                m = re.search(r'(?i)Destination\s*[:：]\s*(.+)', cell_str)
                if m:
                    destination = m.group(1).strip()
                m = re.search(r'(?i)Beneficiary\s+Bank\s*[:：]\s*(.+)', cell_str)
                if m:
                    beneficiary_bank = m.group(1).strip()
            # Port of loading（可能在 payment 区域之外）
            m = re.search(r'(?i)Port\s+of\s+loading\s*[:：]\s*(.+)', cell_str)
            if m:
                loading_port = m.group(1).strip()
            m = re.search(r'(?i)Loading\s+Port\s*[:：]\s*(.+)', cell_str)
            if m:
                loading_port = m.group(1).strip()
            # Invoice To
            m = re.search(r'(?i)INVOICE\s+TO\s*[:.]?\s*NAME\s*[:：]\s*(.+)', cell_str)
            if m:
                invoice_to = m.group(1).strip()
            m = re.search(r'(?i)INVOICE\s+TO\s*[:：]\s*(.+)', cell_str)
            if m:
                invoice_to = m.group(1).strip()

    # ── 7. 提取产品明细 ──────────────────────────────────────────────
    items: list[PiContractItemRow] = []
    table_started = False
    item_index = 0
    pending_desc = None  # 跨行：上一行品名，下一行数字
    table_header_cells = []  # 保存表头行的单元格（用于后续分析）
    hs_code_col = None  # HS CODE 列索引

    for row_idx, row in enumerate(rows):
        row_text = " | ".join(str(c).strip() for c in row)

        # 检测表格开始（DESCRIPTION OF GOODS 或 QUANTITY 表头）
        if re.search(r'(?i)DESCRIPTION\s+(?:OF\s+)?GOODS', row_text):
            table_started = True
            table_header_cells = [str(c).strip() for c in row]
            hs_code_col = _find_hs_code_column(rows[row_idx:row_idx+2])
            continue
        if re.search(r'(?i)QUANTITY', row_text) and not table_started:
            table_started = True
            table_header_cells = [str(c).strip() for c in row]
            hs_code_col = _find_hs_code_column(rows[row_idx:row_idx+2])
            continue
        if not table_started:
            continue

        # 检测表格结束（Payment terms / 总结性行）
        if re.search(r'(?i)Payment\s+terms\s+and\s+conditions', row_text):
            break
        if re.search(r'(?i)AWB\s*#', row_text):
            break
        if re.search(r'(?i)I\s+hereby\s+certify', row_text):
            break

        # 跳过空行
        non_empty = [str(c).strip() for c in row if str(c).strip()]
        if not non_empty:
            pending_desc = None
            continue

        # 跳过 TOTAL 行
        first_cell = str(row[0]).strip() if row[0] else ""
        if re.match(r'(?i)^TOTAL\s*:', first_cell):
            pending_desc = None
            continue
        if re.match(r'(?i)^TOTAL\s+AMOUNT', first_cell):
            pending_desc = None
            continue
        if re.match(r'(?i)^TOTAL\s+(?:US\s+)?DOLLARS', first_cell):
            pending_desc = None
            continue
        if re.match(r'(?i)^TOTAL\s+CNY', first_cell):
            pending_desc = None
            continue
        if re.match(r'(?i)^TOTAL\s+USD', first_cell):
            pending_desc = None
            continue
        if re.match(r'(?i)^100%\s+payment', first_cell):
            pending_desc = None
            continue

        # 跳过 PACKING 说明行
        if _is_packing_row(row_text):
            pending_desc = None
            continue

        # 跳过备注行
        if _is_remark_or_terms_row(row_text):
            pending_desc = None
            continue

        # 跳过数字合计行（如 Row 13 in Minhao: [(4, 20000.0), (6, 19000.0)]）
        if len(non_empty) <= 2 and all(_clean_number(v) is not None for v in non_empty):
            pending_desc = None
            continue

        # 跳过 "双方同意..." / "The undersigned..." 等过渡行
        if re.search(r'(?i)(?:双方同意|The\s+undersigned)', row_text):
            pending_desc = None
            continue

        # ── 解析产品行 ──────────────────────────────────────────
        # 收集所有单元格的值
        cells_data = []
        for ci, cell in enumerate(row):
            val = str(cell).strip()
            cells_data.append((ci, val))

        # 判断这是否是产品行：
        # 条件1: 有文字内容（品名）+ 至少2个数字（数量、金额或单价+金额）
        # 条件2: 有文字内容 + pending_desc（跨行产品）
        has_text = False
        text_content = ""
        numbers = []

        for ci, val in cells_data:
            if not val:
                continue
            cleaned = _clean_number(val)
            if cleaned is not None:
                numbers.append(cleaned)
            elif re.search(r'[A-Za-z\u4e00-\u9fff]', val):
                has_text = True
                if not text_content:
                    text_content = val
                else:
                    text_content += " " + val

        # 产品行需要：至少2个数字（数量+金额 或 单价+金额）
        if len(numbers) >= 2 and has_text:
            # 提取 HS code 从描述中
            hs_in_desc, clean_desc = _extract_hs_from_description(text_content)

            # 从 HS code 列提取（如果表头有 HS CODE 列）
            hs_from_col = None
            if hs_code_col is not None and hs_code_col < len(row):
                col_val = str(row[hs_code_col]).strip()
                if col_val and re.search(r'\d', col_val):
                    hs_from_col = _normalize_hs_code(col_val)

            hs_code = hs_in_desc or hs_from_col or hs_code_from_label

            # 按位置推断数量、单价、金额
            # 通常：数量在 QUANTITY 列，单价在 UNIT PRICE 列，金额在 AMOUNT 列
            quantity = None
            unit_price = None
            amount = None

            # 简单策略：最后一个数字通常是金额，倒数第二个是单价（或数量），倒数第三个是数量
            if len(numbers) == 2:
                quantity = numbers[0]
                amount = numbers[1]
            elif len(numbers) == 3:
                quantity = numbers[0]
                unit_price = numbers[1]
                amount = numbers[2]
            elif len(numbers) >= 4:
                # 跳过可能是 HS code 的纯数字（如 3403910000）
                product_numbers = [n for n in numbers if n < 1000000]
                if len(product_numbers) >= 3:
                    quantity = product_numbers[0]
                    unit_price = product_numbers[1]
                    amount = product_numbers[2]
                elif len(product_numbers) >= 2:
                    quantity = product_numbers[0]
                    amount = product_numbers[1]

            # 构建 notes
            notes_parts = []
            if payment_terms:
                notes_parts.append(f"付款条款: {payment_terms}")
            if beneficiary_bank:
                notes_parts.append(f"受益银行: {beneficiary_bank}")
            if destination:
                notes_parts.append(f"目的地: {destination}")

            # 提取 PACKING 信息（如果产品行后面紧跟 PACKING 行）
            packing_info = None
            if row_idx + 1 < len(rows):
                next_row_text = " | ".join(str(c).strip() for c in rows[row_idx + 1])
                packing_m = re.search(r'(?i)PACKING\s*[:：]\s*(.+)', next_row_text)
                if packing_m:
                    packing_info = packing_m.group(1).strip()

            if packing_info:
                notes_parts.append(f"包装: {packing_info}")

            item_index += 1
            items.append(PiContractItemRow(
                row_index=item_index,
                status="success",
                error_msg=None,
                internal_code=None,
                quantity=quantity,
                unit_price=unit_price,
                total_amount=amount,
                product_color=None,
                hs_code=hs_code,
                customs_name=clean_desc.strip() if clean_desc else text_content.strip(),
                customs_composition=None,
                order_customs_name=None,
                notes="; ".join(notes_parts) if notes_parts else None,
            ))
            pending_desc = None

        elif has_text and not numbers:
            # 可能是跨行产品描述的上半部分
            pending_desc = text_content

        elif pending_desc and numbers:
            # 上一行的品名 + 这一行的数字
            quantity = numbers[0] if len(numbers) >= 1 else None
            amount = numbers[-1] if len(numbers) >= 2 else numbers[0]
            unit_price = numbers[1] if len(numbers) >= 3 else None

            notes_parts = []
            if payment_terms:
                notes_parts.append(f"付款条款: {payment_terms}")
            if destination:
                notes_parts.append(f"目的地: {destination}")

            item_index += 1
            items.append(PiContractItemRow(
                row_index=item_index,
                status="success",
                error_msg=None,
                internal_code=None,
                quantity=quantity,
                unit_price=unit_price,
                total_amount=amount,
                product_color=None,
                hs_code=hs_code_from_label,
                customs_name=pending_desc.strip(),
                customs_composition=None,
                order_customs_name=None,
                notes="; ".join(notes_parts) if notes_parts else None,
            ))
            pending_desc = None
        else:
            # 非产品行，清除 pending_desc
            pending_desc = None

    # ── 8. 构建结果 ──────────────────────────────────────────────────
    if not pi_no:
        raise ValueError("无法识别 PI 编号，请确认文件格式")

    fields_recognized = sum(1 for v in [pi_no, consignee_name, hs_code_from_label, pi_date, destination, price_term] if v)
    total_fields = 6
    confidence = ConfidenceInfo(
        recognized=fields_recognized,
        total=total_fields,
        percentage=f"{fields_recognized / total_fields * 100:.0f}%",
        failed_rows=0,
    )

    return PiContractUploadResponse(
        pi_no=pi_no or "",
        customer_code=None,
        sales_person=None,
        pi_date=pi_date,
        is_ordered="0",
        consignee_name=consignee_name,
        consignee_address=consignee_address,
        consignee_tel=consignee_tel,
        destination=destination,
        loading_port=loading_port,
        price_term=price_term,
        payment_terms=payment_terms,
        invoice_to=invoice_to,
        currency=currency,
        items=items,
        confidence=confidence,
    )


# ──────────────────────────────────────────────────────────────────────
# 文件读取函数
# ──────────────────────────────────────────────────────────────────────

def _read_xlsx_rows(file_path: str) -> list[list[str]]:
    wb = load_workbook(file_path, read_only=False, data_only=True)
    ws = wb.active
    rows: list[list[str]] = []
    for row in ws.iter_rows(values_only=True):
        rows.append([str(cell).strip() if cell is not None else "" for cell in row])
    wb.close()
    return rows


def _read_xls_rows(file_path: str) -> list[list[str]]:
    wb = xlrd.open_workbook(file_path, encoding_override="gbk")
    sheet = wb.sheet_by_index(0)
    rows = []
    for row_idx in range(sheet.nrows):
        row_values = sheet.row_values(row_idx)
        rows.append([str(cell).strip() if cell else "" for cell in row_values])
    wb.release_resources()
    return rows


def _read_xlsx_bytes(content: bytes) -> list[list[str]]:
    wb = load_workbook(io.BytesIO(content), read_only=False, data_only=True)
    ws = wb.active
    rows: list[list[str]] = []
    for row in ws.iter_rows(values_only=True):
        rows.append([str(cell).strip() if cell is not None else "" for cell in row])
    wb.close()
    return rows


def _read_xls_bytes(content: bytes) -> list[list[str]]:
    wb = xlrd.open_workbook(file_contents=content, encoding_override="gbk")
    sheet = wb.sheet_by_index(0)
    rows = []
    for row_idx in range(sheet.nrows):
        row_values = sheet.row_values(row_idx)
        rows.append([str(cell).strip() if cell else "" for cell in row_values])
    wb.release_resources()
    return rows


def _extract_text_from_pdf_bytes(content: bytes) -> str:
    """通过 OCR 从 PDF 内容提取纯文本（图片型扫描 PDF）"""
    pytesseract.pytesseract.tesseract_cmd = os.getenv("TESSERACT_CMD", "/usr/bin/tesseract")
    doc = pymupdf.open(stream=content, filetype="pdf")
    texts = []
    for page in doc:
        mat = pymupdf.Matrix(3, 3)
        pix = page.get_pixmap(matrix=mat)
        img_bytes = pix.tobytes("png")
        img = Image.open(io.BytesIO(img_bytes))
        text = pytesseract.image_to_string(img, lang="chi_sim+eng", config="--psm 6")
        texts.append(text)
    doc.close()
    return "\n".join(texts)


def _parse_rows_from_text(text: str) -> list[list[str]]:
    """将纯文本转换为行列表（用于 Proforma Invoice 解析）"""
    lines = text.split("\n")
    rows = []
    for line in lines:
        stripped = line.strip()
        if not stripped:
            continue
        parts = re.split(r"\s{2,}", stripped)
        if parts and parts[0]:
            rows.append([p.strip() for p in parts])
    return rows


def parse_proforma_invoice_from_text(text: str, filename: str = "") -> PiContractUploadResponse:
    """从 OCR 文本直接解析 Proforma Invoice（用于 PDF）"""

    def scan(label_pat):
        """先找标签同行值，若值看起来像另一标签或为空则检查下一行"""
        idx = re.search(label_pat, text, re.IGNORECASE)
        if not idx:
            return None
        start = idx.end()
        after = text[start:]
        same = after.split(chr(10))[0].strip()
        if not same or re.match(r"^[A-Z]{2,}[:：;]?\s*$", same):
            rest = after.split(chr(10), 2)
            if len(rest) >= 2 and rest[1].strip():
                return rest[1].strip()
            return None
        return same

    def scan_next_line(label_pat):
        """标签在第N行，值在第N+1行"""
        idx = re.search(label_pat, text, re.IGNORECASE)
        if not idx:
            return None
        start = idx.end()
        after = text[start:]
        lines = after.split(chr(10), 2)
        return lines[1].strip() if len(lines) >= 2 and lines[1].strip() else None

    def scan_value(label_pat, value_pat):
        """找标签后, 在后续文本中搜索匹配 value_pat 的值"""
        idx = re.search(label_pat, text, re.IGNORECASE)
        if not idx:
            return None
        window = text[idx.end(): idx.end() + 200]
        m = re.search(value_pat, window)
        return m.group(0).strip() if m else None

    # 1) PI/Order No — 多种 OCR 变形
    pi_no = (
        scan(r"ORDER\s+NO\s*[:;.]?") or
        scan(r"PI\s+NO\s*[:;.]?") or
        scan(r"PI\s+NUMBER\s*[:;.]?") or
        scan(r"ORDER\s+NUMBER\s*[:;.]?") or
        scan_next_line(r"ORDER\s+NO\s*[:;.]?") or
        scan_next_line(r"PI\s+NO\s*[:;.]?") or
        scan_next_line(r"'\s*PI\s+NO\s*;?") or
        (m.group(0) if (m := re.search(r"HT\d{8,10}[A-Z0-9]*", text)) else None)
    )
    if not pi_no and filename:
        fn_pi = re.search(r"HT\d{8,10}[A-Z0-9]*", os.path.basename(filename))
        pi_no = fn_pi.group(0) if fn_pi else None
    if pi_no and len(pi_no) >= 8:
        pass
    else:
        pi_no = None

    # 2) consignee NAME
    consignee_name = scan(r"NAME\s*:")
    if not consignee_name:
        m = re.search(r"(?mi)^NAME\s*:\s*(.+)$", text, re.MULTILINE)
        consignee_name = m.group(1).strip() if m else None

    # 3) consignee ADD
    consignee_address = scan(r"ADD\s*:") or scan(r"ADDRESS\s*:")
    if consignee_address:
        consignee_address = re.sub(r"\s+\d{4}[-/.]\d{1,2}[-/.]\d{1,2}\s*$", "", consignee_address).strip()

    # 3b) consignee TEL
    consignee_tel = scan(r"TEL\s*:") or scan(r"Tel\s*:") or scan(r"Telephone\s*:")

    # 4) destination
    destination = scan(r"[0-9]\.\s*DEST\s*:") or scan(r"DEST\s*INAT\s*ION\s*:")

    # 5) date
    pi_date = scan_value(r"DATE\s*[:;]?", r"\d{4}[-/.]\d{1,2}[-/.]\d{1,2}") or scan(r"DATE\s*:")

    # 6) Port of loading
    loading_port = scan(r"Port\s+of\s+loading\s*:") or scan(r"Loading\s+Port\s*:")

    # 7) Price Term
    price_term = scan(r"Price\s+Term\s*:")
    if not price_term:
        m = re.search(r"\b(CIF|FOB|CFR|C\s*&\s*F|EXW|FCA|CPT|CIP|DAP|DPU)\b", text)
        price_term = m.group(0) if m else None

    # 8) Invoice To
    invoice_to = scan(r"INVOICE\s+TO\s*[:.]?\s*NAME\s*:") or scan(r"INVOICE\s+TO\s*:")

    # 产品明细行解析
    items = []
    lines = text.split(chr(10))
    started = False
    idx2 = 0
    pending_desc = None
    for line in lines:
        s = line.strip()
        if not s:
            continue
        if "DESCRIPTION" in s.upper() or "QUANTITY" in s.upper():
            started = True
            continue
        if not started:
            continue
        if s.startswith("1.") and "Packing" in s or s.startswith("TOTAL") or "Payment" in s:
            break

        m = re.match(r"(.+?)\s+([\d,]+\.?\d*)\s+\|?\s*([\d,]+\.?\d*)\s+\|?\s*US?\\\$?\s+([\d,]+\.?\d*)", s)
        if m:
            desc, qty_raw, price_raw, amount_raw = m.groups()
        else:
            num_m = re.match(r"^([\d,]+\.?\d*)\s+([\d,]+\.?\d*)$", s)
            if num_m and pending_desc:
                qty_raw, amount_raw = num_m.groups()
                price_raw = None
                desc = pending_desc
            else:
                if not re.match(r"^[\d,. ]+$", s):
                    pending_desc = s
                continue

        try:
            qty_val = float(qty_raw.replace(",", "")) if qty_raw else None
            price_val = float(price_raw.replace(",", "")) if price_raw else None
            amount_val = float(amount_raw.replace(",", "")) if amount_raw else None
        except ValueError:
            qty_val = price_val = amount_val = None
        idx2 += 1
        notes = []
        if destination:
            notes.append(f"目的地: {destination}")
        items.append(PiContractItemRow(
            row_index=idx2,
            status="success",
            error_msg=None,
            internal_code=None,
            quantity=qty_val,
            unit_price=price_val,
            total_amount=amount_val,
            product_color=None,
            hs_code=None,
            customs_name=desc.strip() if desc else s,
            customs_composition=None,
            order_customs_name=None,
            notes="; ".join(notes) if notes else None,
        ))
        pending_desc = None

    if not pi_no:
        raise ValueError("无法识别 PI 编号，请确认文件格式")

    recognized = sum(1 for v in [pi_no] if v)
    confidence = ConfidenceInfo(
        recognized=recognized,
        total=1,
        percentage="100%" if recognized else "0%",
        failed_rows=0,
    )

    return PiContractUploadResponse(
        pi_no=pi_no or "",
        customer_code=None,
        sales_person=None,
        pi_date=pi_date,
        is_ordered="0",
        consignee_name=consignee_name,
        consignee_address=consignee_address,
        consignee_tel=consignee_tel,
        destination=destination,
        loading_port=loading_port,
        price_term=price_term,
        invoice_to=invoice_to,
        items=items,
        confidence=confidence,
    )


# ──────────────────────────────────────────────────────────────────────
# DeepSeek AI 解析
# ──────────────────────────────────────────────────────────────────────

def _format_rows_for_ai(rows: list[list[str]]) -> str:
    """将行列数据格式化为AI可读的结构化文本"""
    lines = []
    for row_idx, row in enumerate(rows):
        cells = []
        for col_idx, cell in enumerate(row):
            if cell.strip():
                cells.append(f"[R{row_idx+1}C{col_idx+1}]{cell}")
        if cells:
            lines.append(" | ".join(cells))
    return "\n".join(lines)


def _safe_float(value) -> float | None:
    """安全转换为浮点数"""
    if value is None:
        return None
    try:
        return float(value)
    except (ValueError, TypeError):
        return None


def _build_response_from_ai(ai_data: dict) -> PiContractUploadResponse:
    """将AI返回的JSON转换为PiContractUploadResponse"""
    items = []
    for idx, item in enumerate(ai_data.get("items", [])):
        items.append(PiContractItemRow(
            row_index=idx + 1,
            status="success",
            error_msg=None,
            internal_code=item.get("internal_code"),
            quantity=_safe_float(item.get("quantity")),
            unit_price=_safe_float(item.get("unit_price")),
            total_amount=_safe_float(item.get("total_amount")),
            product_color=item.get("product_color"),
            hs_code=item.get("hs_code"),
            customs_name=item.get("customs_name"),
            customs_composition=item.get("customs_composition"),
            order_customs_name=None,
            notes=item.get("notes"),
        ))

    # 计算置信度（基于识别到的字段数）
    fields = ["pi_no", "customer_code", "consignee_name", "destination", "price_term", "currency"]
    recognized = sum(1 for f in fields if ai_data.get(f))
    confidence = ConfidenceInfo(
        recognized=recognized,
        total=len(fields),
        percentage=f"{recognized / len(fields) * 100:.0f}%",
        failed_rows=0,
    )

    return PiContractUploadResponse(
        pi_no=ai_data.get("pi_no") or "",
        customer_code=ai_data.get("customer_code"),
        sales_person=ai_data.get("sales_person"),
        pi_date=ai_data.get("pi_date"),
        is_ordered="0",
        consignee_name=ai_data.get("consignee_name"),
        consignee_address=ai_data.get("consignee_address"),
        consignee_tel=ai_data.get("consignee_tel"),
        destination=ai_data.get("destination"),
        loading_port=ai_data.get("loading_port"),
        price_term=ai_data.get("price_term"),
        payment_terms=ai_data.get("payment_terms"),
        invoice_to=ai_data.get("invoice_to"),
        currency=ai_data.get("currency"),
        items=items,
        confidence=confidence,
    )


def _call_deepseek_api(rows: list[list[str]], filename: str) -> PiContractUploadResponse | None:
    """
    调用DeepSeek-V4-Flash API解析PI文件
    返回解析结果或None（失败时）
    """
    from app.core.config import DEEPSEEK_API_KEY

    if not DEEPSEEK_API_KEY:
        print("[AI Parse] DEEPSEEK_API_KEY未配置，跳过AI解析")
        return None

    # 格式化Excel内容
    content = _format_rows_for_ai(rows)

    system_prompt = """你是一个PI(Proforma Invoice)合同解析专家。请从Excel内容中提取以下字段，返回JSON格式。

字段列表：
- pi_no: PI号/合同号（必须包含HT开头的编号）
- customer_code: 客户编码
- sales_person: 业务员
- pi_date: 日期
- consignee_name: 收货人名称
- consignee_address: 收货人地址
- consignee_tel: 收货人电话
- destination: 目的港
- loading_port: 装货港
- price_term: 价格条款（FOB/CIF/CFR等）
- payment_terms: 付款条款
- currency: 币制（USD/CNY/RMB）
- invoice_to: 发票抬头
- items: 产品明细数组，每项包含：
  - internal_code: 内部编码/产品代码
  - quantity: 数量（数字）
  - unit_price: 单价（数字）
  - total_amount: 金额（数字）
  - product_color: 产品颜色
  - hs_code: 海关编码/H.S.Code
  - customs_name: 报关品名/产品描述
  - customs_composition: 报关成分
  - notes: 备注

JSON输出格式示例：
{
  "pi_no": "HT20260315A",
  "customer_code": "TOA-DOVECHEM",
  "sales_person": "张三",
  "pi_date": "2026-03-15",
  "consignee_name": "PT ABC Indonesia",
  "consignee_address": "Jl. Sudirman No.123, Jakarta",
  "consignee_tel": "+62-21-1234567",
  "destination": "Jakarta, Indonesia",
  "loading_port": "Shanghai",
  "price_term": "CIF Jakarta",
  "payment_terms": "T/T 30 days after B/L date",
  "currency": "USD",
  "invoice_to": "PT ABC Indonesia",
  "items": [
    {
      "internal_code": "SILI-001",
      "quantity": 2400,
      "unit_price": 29.5,
      "total_amount": 70800,
      "product_color": "White",
      "hs_code": "3402.90.99",
      "customs_name": "FIXING HT-7212",
      "customs_composition": "Active matter: 15%",
      "notes": ""
    }
  ]
}

注意：
1. 无法识别的字段填null
2. 数值字段保持数字类型，不要加引号
3. items数组包含所有产品明细行
4. 如果有多个产品，全部放入items数组
5. 严格按照上述JSON格式返回，不要添加额外文字"""

    try:
        from openai import OpenAI

        client = OpenAI(
            api_key=DEEPSEEK_API_KEY,
            base_url="https://api.deepseek.com"
        )

        response = client.chat.completions.create(
            model="deepseek-v4-flash",
            messages=[
                {"role": "system", "content": system_prompt},
                {"role": "user", "content": f"文件名: {filename}\n\nExcel内容:\n{content}"}
            ],
            max_tokens=4096,
            response_format={'type': 'json_object'},
            extra_body={"thinking": {"type": "disabled"}}
        )

        ai_content = response.choices[0].message.content

        # 处理空响应
        if not ai_content:
            print("[AI Parse] AI返回空内容")
            return None

        # 解析JSON
        ai_data = json.loads(ai_content)
        return _build_response_from_ai(ai_data)

    except Exception as e:
        print(f"[AI Parse] DeepSeek API调用失败: {e}")
        return None


def _auto_parse_rows(rows: list[list[str]]) -> PiContractUploadResponse:
    """
    自动识别格式并解析：先尝试表格解析，失败则尝试 Proforma Invoice 格式。
    """
    try:
        return parse_pi_file(rows)
    except ValueError as e:
        if "缺少关键字段" in str(e) or "表头行无效" in str(e) or "无法从数据中提取 PI号" in str(e):
            return parse_proforma_invoice(rows)
        raise


def parse_file_path(file_path: str) -> PiContractUploadResponse:
    """解析PI文件路径，支持 .xlsx 和 .xls 格式"""
    ext = os.path.splitext(file_path)[1].lower()

    if ext == ".xlsx":
        rows = _read_xlsx_rows(file_path)
    elif ext == ".xls":
        rows = _read_xls_rows(file_path)
    else:
        raise ValueError(f"不支持的文件格式: {ext}，仅支持 .xlsx 和 .xls")

    return _auto_parse_rows(rows)


def parse_pi_bytes(content: bytes, filename: str) -> PiContractUploadResponse:
    """从内存字节解析PI文件，AI优先，Regex兜底，支持 .xlsx / .xls / .pdf"""
    ext = os.path.splitext(filename)[1].lower() if filename else ""

    if ext == ".xlsx":
        rows = _read_xlsx_bytes(content)
        # 1. 尝试AI解析（主路径）
        ai_result = _call_deepseek_api(rows, filename)
        if ai_result:
            print(f"[AI Parse] 成功解析PI: {ai_result.pi_no}, {len(ai_result.items)}个产品")
            return ai_result
        # 2. AI失败，fallback到Regex
        print("[AI Parse] AI解析失败，使用Regex兜底")
        return _auto_parse_rows(rows)
    elif ext == ".xls":
        rows = _read_xls_bytes(content)
        # 1. 尝试AI解析（主路径）
        ai_result = _call_deepseek_api(rows, filename)
        if ai_result:
            print(f"[AI Parse] 成功解析PI: {ai_result.pi_no}, {len(ai_result.items)}个产品")
            return ai_result
        # 2. AI失败，fallback到Regex
        print("[AI Parse] AI解析失败，使用Regex兜底")
        return _auto_parse_rows(rows)
    elif ext == ".pdf":
        text = _extract_text_from_pdf_bytes(content)
        return parse_proforma_invoice_from_text(text, filename)
    else:
        raise ValueError(f"不支持的文件格式: {ext}，仅支持 .xlsx、.xls 和 .pdf")

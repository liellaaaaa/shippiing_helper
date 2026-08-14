# backend/tests/test_customs_dynamic_rows.py
"""报关单 sheet 按产品数动态扩展的测试"""
import io
import openpyxl
import pytest

from app.services.document_service import DocumentService
from app.schemas.ledger import LedgerRecordResponse, LedgerItemSchema


def make_item(i: int) -> LedgerItemSchema:
    """构造第 i 个产品（1-based）"""
    return LedgerItemSchema(
        internal_code=f"SILI-{i:03d}",
        product_cn=f"有机硅柔软剂{i}",
        quantity_kg=1000.0 * i,
        unit_price=10.0,
        total_amount=10000.0 * i,
        hs_code="3910000000",
        customs_name=f"有机硅柔软剂{i}",
        customs_ingredients="八甲基环四硅氧烷50%,水50%",
        drum_count=2 * i,
        pallet_count=i,
        gross_weight_kg=1100.0 * i,
        net_weight_kg=1000.0 * i,
    )


def make_record(n_items: int) -> LedgerRecordResponse:
    return LedgerRecordResponse(
        id=1,
        order_no="HT260829E01",
        customer_code="TOA-DOVECHEM",
        sales_person="张三",
        consignee_name="ABC Co., Ltd.",
        consignee_address="123 Main St, Mumbai",
        destination="Nhava Sheva, India",
        loading_port="Shenzhen",
        price_term="CIF",
        payment_terms="T/T",
        pi_date="2026-08-01",
        currency="USD",
        items=[make_item(i) for i in range(1, n_items + 1)],
        status="saved",
    )


def generate_customs_bytes(monkeypatch, n_items: int) -> bytes:
    """调用 generate_customs 并返回生成的 xlsx 字节"""
    from app.services import ledger_service
    monkeypatch.setattr(
        ledger_service.LedgerService,
        "get_ledger_record",
        lambda self, record_id: make_record(n_items),
    )
    svc = DocumentService()
    content, _doc_key, _b64 = svc.generate_customs(ledger_record_id=1)
    return content


def load_customs_sheet(monkeypatch, n_items: int):
    wb = openpyxl.load_workbook(io.BytesIO(generate_customs_bytes(monkeypatch, n_items)))
    return wb["报关单"], wb


def test_n6_unchanged(monkeypatch):
    """N=6：模板预建容量内，不触发扩展，max_row 保持 37"""
    ws, _ = load_customs_sheet(monkeypatch, 6)
    assert ws.max_row == 37
    # 打印区域不变
    assert ws.print_area == "'报关单'!$A$1:$S$37"


def test_n7_adds_one_block(monkeypatch):
    """N=7：多出 1 块（3 行），max_row 变为 40"""
    ws, _ = load_customs_sheet(monkeypatch, 7)
    assert ws.max_row == 40
    assert ws.print_area == "'报关单'!$A$1:$S$40"
    # 第 7 个产品数据写入：项号=7, HS Code, 数量, 币制行
    assert ws.cell(38, 1).value == 7          # A38: 项号
    assert ws.cell(38, 2).value == "3910000000"  # B38: 商品编号
    assert ws.cell(38, 7).value == 7000.0     # G38: 数量
    assert ws.cell(40, 7).value == 7000.0     # G40: quantity reference
    assert ws.cell(40, 9).value == "美元"      # I40: 币制中文


def test_n17_many_blocks(monkeypatch):
    """N=17：共 17 块，max_row = 20 + 3*17 - 1 = 70"""
    ws, _ = load_customs_sheet(monkeypatch, 17)
    assert ws.max_row == 70
    assert ws.print_area == "'报关单'!$A$1:$S$70"
    assert ws.cell(68, 1).value == 17         # 最后一块数据行（20+16*3=68）
    assert ws.cell(68, 7).value == 17000.0


def test_extended_block_borders_and_merges(monkeypatch):
    """新块样式：边框与合并单元格必须与源块（第 6 块）一致"""
    ws, _ = load_customs_sheet(monkeypatch, 8)
    # 源块第 6 块：35/36/37 行；新块第 7 块：38/39/40 行
    # 边框抽样：数据行 A 列、申报行 A 列必须有 thin 边框
    assert ws.cell(38, 1).border.left.style is not None   # A38 边框
    assert ws.cell(39, 1).border.left.style is not None   # A39 申报要素行边框
    # 合并单元格：新块申报要素合并跨两行 D39:F40（对应源块 D36:F37）
    assert "D39:F40" in [str(r) for r in ws.merged_cells.ranges]
    # 数据行合并存在：B38:C38
    assert "B38:C38" in [str(r) for r in ws.merged_cells.ranges]


def test_n5_regression(monkeypatch):
    """N=5：小于预建容量，行为与扩展无关，数据正确写入"""
    ws, _ = load_customs_sheet(monkeypatch, 5)
    assert ws.max_row == 37
    assert ws.cell(20, 1).value == 1
    assert ws.cell(32, 1).value == 5


def test_invoice_contract_row_alignment(monkeypatch):
    """发票页产品连续填充（行8起），与合同页公式引用对齐"""
    wb = openpyxl.load_workbook(io.BytesIO(generate_customs_bytes(monkeypatch, 8)))
    ws_inv = wb["发票"]
    # 产品连续填充：行8-15
    assert ws_inv.cell(8, 3).value == "有机硅柔软剂1"
    assert ws_inv.cell(11, 3).value == "有机硅柔软剂4"
    assert ws_inv.cell(12, 3).value == "有机硅柔软剂5"
    assert ws_inv.cell(13, 3).value == "有机硅柔软剂6"
    assert ws_inv.cell(14, 3).value == "有机硅柔软剂7"
    assert ws_inv.cell(15, 3).value == "有机硅柔软剂8"
    # 汇总行在行24（N=8 <= 16，不触发扩展）
    assert ws_inv.cell(24, 8).value is not None


# ── 发票 Sheet 动态扩展测试 ──────────────────────────────────────────

def load_invoice_sheet(monkeypatch, n_items: int):
    """获取生成后的发票 sheet"""
    wb = openpyxl.load_workbook(io.BytesIO(generate_customs_bytes(monkeypatch, n_items)))
    return wb["发票"], wb


def test_invoice_n10_no_extension(monkeypatch):
    """N=10（< 预建16）：不触发扩展，产品行 8-17，汇总行 24"""
    ws, _ = load_invoice_sheet(monkeypatch, 10)
    assert ws.cell(8, 3).value == "有机硅柔软剂1"
    assert ws.cell(17, 3).value == "有机硅柔软剂10"
    # 汇总行仍在行24
    assert ws.cell(24, 8).value is not None
    assert ws.cell(24, 8).value == 550000.0  # 10000 * (1+2+...+10) = 10000*55
    # 打印区域不变
    assert ws.print_area == "'发票'!$A$1:$H$27"


def test_invoice_n16_full_prebuilt(monkeypatch):
    """N=16（= 预建容量）：不触发扩展，产品行 8-23，汇总行 24"""
    ws, _ = load_invoice_sheet(monkeypatch, 16)
    assert ws.cell(8, 3).value == "有机硅柔软剂1"
    assert ws.cell(23, 3).value == "有机硅柔软剂16"
    # 汇总行仍在行24
    assert ws.cell(24, 8).value is not None
    # 打印区域不变
    assert ws.print_area == "'发票'!$A$1:$H$27"


def test_invoice_n20_extend_4_rows(monkeypatch):
    """N=20（扩展4行）：产品行 8-27，汇总行下移到 28"""
    ws, _ = load_invoice_sheet(monkeypatch, 20)
    # 产品行范围
    assert ws.cell(8, 3).value == "有机硅柔软剂1"
    assert ws.cell(27, 3).value == "有机硅柔软剂20"
    # 汇总行下移到 28 = 24 + (20-16)
    assert ws.cell(28, 8).value is not None   # H: 数字金额
    assert ws.cell(28, 3).value is not None   # C: 中文大写金额
    assert ws.cell(28, 7).value is not None   # G: 币制
    # 打印区域扩展 4 行（原 27 + 4 = 31）
    assert ws.print_area == "'发票'!$A$1:$H$31"
    # 新行样式：行 24 应有行高（从行 23 复制）
    assert ws.row_dimensions[24].height == 30.0
    # 新行样式：A 列有边框（从行 23 复制）
    assert ws.cell(24, 1).border.left.style is not None


def test_invoice_n38_extend_22_rows(monkeypatch):
    """N=38（扩展22行）：产品行 8-45，汇总行下移到 46"""
    ws, _ = load_invoice_sheet(monkeypatch, 38)
    # 产品行范围
    assert ws.cell(8, 3).value == "有机硅柔软剂1"
    assert ws.cell(45, 3).value == "有机硅柔软剂38"
    # 汇总行下移到 46 = 24 + (38-16)
    assert ws.cell(46, 8).value is not None
    assert ws.cell(46, 3).value is not None
    assert ws.cell(46, 7).value is not None
    # 打印区域扩展 22 行（原 27 + 22 = 49）
    assert ws.print_area == "'发票'!$A$1:$H$49"
    # 新行样式验证
    assert ws.row_dimensions[24].height == 30.0
    assert ws.row_dimensions[45].height == 30.0
    assert ws.cell(24, 1).border.left.style is not None
    assert ws.cell(45, 1).border.left.style is not None


# ── 箱单 Sheet 动态扩展测试 ──────────────────────────────────────────

def load_packing_sheet(monkeypatch, n_items: int):
    wb = openpyxl.load_workbook(io.BytesIO(generate_customs_bytes(monkeypatch, n_items)))
    return wb["箱单"], wb


def test_packing_n10_no_extend(monkeypatch):
    """N=10（< 预建16行）：不触发扩展，汇总行保持在行26"""
    ws, _ = load_packing_sheet(monkeypatch, 10)
    # 汇总行仍在行26
    assert ws.cell(26, 3).value == "=SUM(C10:C25)"
    # 打印区域不变
    assert ws.print_area == "'箱单'!$A$1:$H$28"
    # 产品数据正确写入
    assert ws.cell(10, 2).value == "有机硅柔软剂1"
    assert ws.cell(19, 2).value == "有机硅柔软剂10"


def test_packing_n16_exact(monkeypatch):
    """N=16（= 预建容量）：不触发扩展，行25为最后一个产品行"""
    ws, _ = load_packing_sheet(monkeypatch, 16)
    assert ws.cell(26, 3).value == "=SUM(C10:C25)"
    assert ws.print_area == "'箱单'!$A$1:$H$28"
    assert ws.cell(25, 2).value == "有机硅柔软剂16"


def test_packing_n20_extend4(monkeypatch):
    """N=20（扩展4行）：汇总行下移到行30，打印区域更新"""
    ws, _ = load_packing_sheet(monkeypatch, 20)
    # 汇总行下移到 26+4=30
    assert ws.cell(30, 3).value == "=SUM(C10:C29)"
    assert ws.cell(30, 5).value == "=SUM(E10:E29)"
    assert ws.cell(30, 7).value == "=SUM(G10:G29)"
    assert ws.cell(30, 8).value == "=SUM(H10:H29)"
    # 打印区域更新（汇总30 + 页脚2行 = 32）
    assert ws.print_area == "'箱单'!$A$1:$H$32"
    # 扩展行有样式（A列左边框，与模板行25一致）
    assert ws.cell(26, 1).border.left.style == "thin"
    # 扩展行有正确行高
    assert ws.row_dimensions[26].height == 30.0
    # 最后一个产品在行29
    assert ws.cell(29, 2).value == "有机硅柔软剂20"


def test_packing_n38_extend22(monkeypatch):
    """N=38（扩展22行）：汇总行下移到行48，所有产品正确填充"""
    ws, _ = load_packing_sheet(monkeypatch, 38)
    # 汇总行下移到 26+22=48
    assert ws.cell(48, 3).value == "=SUM(C10:C47)"
    assert ws.cell(48, 5).value == "=SUM(E10:E47)"
    assert ws.cell(48, 7).value == "=SUM(G10:G47)"
    assert ws.cell(48, 8).value == "=SUM(H10:H47)"
    # 打印区域（汇总48 + 页脚2行 = 50）
    assert ws.print_area == "'箱单'!$A$1:$H$50"
    # 第一个和最后一个产品
    assert ws.cell(10, 2).value == "有机硅柔软剂1"
    assert ws.cell(47, 2).value == "有机硅柔软剂38"
    # 扩展行样式（A列左边框）
    assert ws.cell(30, 1).border.left.style == "thin"
    assert ws.row_dimensions[30].height == 30.0
